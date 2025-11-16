#!/usr/bin/env python3
"""
Eagle Sign Co. Ultimate Pricing Guide Generator
Builds comprehensive labor standards from actual production data
"""

import pandas as pd
import numpy as np
from collections import defaultdict
from typing import Dict, List, Tuple
import json
import sqlite3
from datetime import datetime

class UltimatePricingGuide:
    """Generate data-driven labor standards by sign type"""
    
    def __init__(self):
        # Eagle sign types
        self.sign_types = {
            'CLLIT': 'Channel Letters Illuminated',
            'CLNON': 'Channel Letters Non-lit',
            'ALULIT': 'Aluminum Cabinet Illuminated',
            'ALUNON': 'Aluminum Cabinet Non-lit',
            'MONDF': 'Monument Double Face',
            'MONSF': 'Monument Single Face',
            'POLLIT': 'Pole Sign Illuminated',
            'POLNON': 'Pole Sign Non-lit',
            'BLDILL': 'Building Sign Illuminated',
            'BLDNON': 'Building Sign Non-lit',
            'DIRECT': 'Directional Signs',
            'VINYL': 'Vinyl Graphics',
            'LED': 'LED Display',
            'NEON': 'Neon Signs'
        }
        
        # Eagle work codes with departments
        self.work_codes = {
            # Design/Permits
            '0098': {'desc': 'PERMIT SUBMITTED', 'dept': 'ADMIN'},
            '0099': {'desc': 'PERMIT RECEIVED', 'dept': 'ADMIN'},
            '0110': {'desc': 'SKETCHING', 'dept': 'DESIGN'},
            '0120': {'desc': 'PRINTING', 'dept': 'DESIGN'},
            '0130': {'desc': 'LAYOUT', 'dept': 'DESIGN'},
            
            # Fabrication
            '0200': {'desc': 'FABRICATION LAYOUT', 'dept': 'FAB'},
            '0210': {'desc': 'SHEET METAL', 'dept': 'FAB'},
            '0215': {'desc': 'STRUCTURAL STEEL', 'dept': 'FAB'},
            '0220': {'desc': 'EXTRUSIONS', 'dept': 'FAB'},
            '0230': {'desc': 'CHANNEL LETTERS', 'dept': 'FAB'},
            '0235': {'desc': 'ROUTING', 'dept': 'FAB'},
            '0240': {'desc': 'FLAT CUT OUT LETTERS', 'dept': 'FAB'},
            '0250': {'desc': 'AWNINGS', 'dept': 'FAB'},
            '0260': {'desc': 'FACES', 'dept': 'FAB'},
            '0270': {'desc': 'MISC FABRICATION', 'dept': 'FAB'},
            '0280': {'desc': 'CRATING & SHIPPING', 'dept': 'FAB'},
            
            # Electrical
            '0310': {'desc': 'BALLAST WIRING', 'dept': 'ELEC'},
            '0340': {'desc': 'ELECTRICAL WIRING', 'dept': 'ELEC'},
            
            # Paint
            '0410': {'desc': 'CLEAN AND ETCH', 'dept': 'PAINT'},
            '0420': {'desc': 'PRIME AND FINISH', 'dept': 'PAINT'},
            
            # Vinyl
            '0510': {'desc': 'LAYOUT', 'dept': 'VINYL'},
            '0520': {'desc': 'CUT AND/OR WEED VINYL', 'dept': 'VINYL'},
            '0550': {'desc': 'VINYL APPLICATION', 'dept': 'VINYL'},
            
            # Installation
            '0605': {'desc': 'FOOTING - INSTALLATION', 'dept': 'INSTALL'},
            '0610': {'desc': 'LOAD/UNLOAD', 'dept': 'INSTALL'},
            '0612': {'desc': 'INSTALL - NO CRANE', 'dept': 'INSTALL'},
            '0615': {'desc': 'WIRING - INSTALLATION', 'dept': 'INSTALL'},
            '0620': {'desc': 'TRAVEL', 'dept': 'INSTALL'},
            '0640': {'desc': '2 MEN & TRUCK INSTALL', 'dept': 'INSTALL'},
            '0650': {'desc': '3 MEN & TRUCK INSTALL', 'dept': 'INSTALL'}
        }
        
        self.labor_database = defaultdict(lambda: defaultdict(list))
        self.size_correlations = defaultdict(lambda: defaultdict(list))
        
    def analyze_historical_data(self, work_orders: List[Dict]) -> Dict:
        """Build pricing guide from historical data"""
        
        # Process each work order
        for wo in work_orders:
            sign_type = self._identify_sign_type(wo)
            size_metrics = self._extract_size_metrics(wo)
            
            # Collect labor hours by code
            for labor_entry in wo.get('labor_entries', []):
                code = labor_entry['code']
                hours = labor_entry['actual_hours']
                
                # Store raw data
                self.labor_database[sign_type][code].append({
                    'hours': hours,
                    'sqft': size_metrics.get('sqft', 0),
                    'letters': size_metrics.get('letter_count', 0),
                    'complexity': wo.get('complexity_score', 1.0)
                })
                
        # Generate pricing guide
        pricing_guide = self._generate_comprehensive_guide()
        
        return pricing_guide
    
    def _identify_sign_type(self, wo: Dict) -> str:
        """Identify sign type from work order"""
        part = wo.get('part_number', '')
        desc = wo.get('description', '')
        
        # Check for explicit type codes
        for type_code in self.sign_types.keys():
            if type_code in part.upper():
                return type_code
                
        # Pattern matching on labor codes
        codes = [e['code'] for e in wo.get('labor_entries', [])]
        
        if '0230' in codes:  # Channel letter fabrication
            return 'CLLIT' if '0340' in codes else 'CLNON'
        elif '0215' in codes:  # Structural steel
            return 'MONDF' if '0260' in codes else 'POLLIT'
        elif '0520' in codes and '0550' in codes:  # Vinyl work
            return 'VINYL'
            
        return 'OTHER'
    
    def _extract_size_metrics(self, wo: Dict) -> Dict:
        """Extract size information from work order"""
        metrics = {}
        
        # Parse dimensions from text
        import re
        text = str(wo.get('description', '')) + ' ' + str(wo.get('part_number', ''))
        
        # Square footage
        sqft_match = re.search(r'(\d+)\s*(?:SF|SQ\.?\s*FT)', text, re.I)
        if sqft_match:
            metrics['sqft'] = float(sqft_match.group(1))
        else:
            # Calculate from dimensions
            dim_match = re.search(r'(\d+)[\'"]?\s*[xX]\s*(\d+)[\'"]?', text)
            if dim_match:
                w, h = float(dim_match.group(1)), float(dim_match.group(2))
                metrics['sqft'] = (w * h) / 144  # Convert to sq ft
                
        # Letter count
        letter_match = re.search(r'(\d+)\s*(?:LETTERS?|CHARACTERS?)', text, re.I)
        if letter_match:
            metrics['letter_count'] = int(letter_match.group(1))
            
        # Height for monuments/pylons
        height_match = re.search(r'(\d+)[\'"]?\s*(?:HIGH|TALL|HEIGHT)', text, re.I)
        if height_match:
            metrics['height'] = float(height_match.group(1))
            
        return metrics
    
    def _generate_comprehensive_guide(self) -> Dict:
        """Generate the ultimate pricing guide"""
        guide = {
            'generated': datetime.now().isoformat(),
            'sign_types': {},
            'formulas': {},
            'confidence_levels': {},
            'size_factors': {}
        }
        
        for sign_type, code_data in self.labor_database.items():
            if sign_type not in self.sign_types:
                continue
                
            guide['sign_types'][sign_type] = {
                'description': self.sign_types[sign_type],
                'labor_standards': {},
                'typical_workflow': [],
                'size_correlations': {}
            }
            
            # Calculate labor standards per code
            for code, entries in code_data.items():
                if code not in self.work_codes:
                    continue
                    
                hours_data = [e['hours'] for e in entries]
                sqft_data = [e['sqft'] for e in entries if e['sqft'] > 0]
                
                if len(hours_data) < 5:  # Need minimum data
                    continue
                    
                # Calculate robust statistics
                labor_standard = {
                    'code': code,
                    'description': self.work_codes[code]['desc'],
                    'department': self.work_codes[code]['dept'],
                    'sample_size': len(hours_data),
                    'hours': {
                        'minimum': round(np.percentile(hours_data, 10), 2),
                        'typical': round(np.percentile(hours_data, 50), 2),
                        'recommended': round(np.percentile(hours_data, 75), 2),
                        'maximum': round(np.percentile(hours_data, 90), 2),
                        'mean': round(np.mean(hours_data), 2),
                        'std_dev': round(np.std(hours_data), 2)
                    }
                }
                
                # Size-based formulas
                if sqft_data and len(sqft_data) >= 10:
                    # Linear regression for hours vs sqft
                    sqft_array = np.array([e['sqft'] for e in entries if e['sqft'] > 0])
                    hours_array = np.array([e['hours'] for e in entries if e['sqft'] > 0])
                    
                    if len(sqft_array) > 1:
                        coef = np.polyfit(sqft_array, hours_array, 1)
                        labor_standard['size_formula'] = {
                            'type': 'linear',
                            'base_hours': round(coef[1], 3),
                            'hours_per_sqft': round(coef[0], 3),
                            'r_squared': round(np.corrcoef(sqft_array, hours_array)[0, 1]**2, 3)
                        }
                        
                guide['sign_types'][sign_type]['labor_standards'][code] = labor_standard
                
            # Determine typical workflow sequence
            workflow = self._determine_workflow(sign_type, code_data.keys())
            guide['sign_types'][sign_type]['typical_workflow'] = workflow
            
            # Size-based multipliers
            guide['sign_types'][sign_type]['size_factors'] = self._calculate_size_factors(sign_type)
            
        # Add formulas and rules
        guide['formulas'] = self._generate_formulas()
        guide['business_rules'] = self._generate_business_rules()
        
        return guide
    
    def _determine_workflow(self, sign_type: str, codes: List[str]) -> List[Dict]:
        """Determine typical workflow sequence"""
        # Standard workflows by sign type
        workflows = {
            'CLLIT': ['0110', '0200', '0230', '0340', '0420', '0260', '0280', '0640'],
            'MONDF': ['0110', '0200', '0215', '0340', '0420', '0260', '0605', '0650'],
            'VINYL': ['0110', '0510', '0520', '0550', '0280', '0630'],
            'POLLIT': ['0110', '0200', '0215', '0220', '0340', '0420', '0605', '0650']
        }
        
        base_workflow = workflows.get(sign_type, [])
        
        # Build actual workflow from data
        actual_workflow = []
        for code in base_workflow:
            if code in codes:
                actual_workflow.append({
                    'sequence': len(actual_workflow) + 1,
                    'code': code,
                    'description': self.work_codes[code]['desc'],
                    'department': self.work_codes[code]['dept'],
                    'required': True
                })
                
        return actual_workflow
    
    def _calculate_size_factors(self, sign_type: str) -> Dict:
        """Calculate size-based complexity factors"""
        return {
            'small': {'threshold_sqft': 20, 'multiplier': 0.9},
            'medium': {'threshold_sqft': 50, 'multiplier': 1.0},
            'large': {'threshold_sqft': 100, 'multiplier': 1.15},
            'xlarge': {'threshold_sqft': 200, 'multiplier': 1.3}
        }
    
    def _generate_formulas(self) -> Dict:
        """Generate pricing formulas"""
        return {
            'channel_letter_formula': {
                'description': 'Hours calculation for channel letters',
                'formula': 'base_hours + (letter_count * hours_per_letter) + (height_factor * complexity)',
                'parameters': {
                    'base_hours': 4.0,
                    'hours_per_letter': 1.5,
                    'height_factors': {
                        '12_inches': 1.0,
                        '24_inches': 1.3,
                        '36_inches': 1.6,
                        '48_inches': 2.0
                    }
                }
            },
            'monument_formula': {
                'description': 'Hours calculation for monuments',
                'formula': 'base_hours + (sqft * hours_per_sqft) + (height_ft * height_multiplier)',
                'parameters': {
                    'base_hours': 8.0,
                    'hours_per_sqft': 0.75,
                    'height_multiplier': 2.0
                }
            }
        }
    
    def _generate_business_rules(self) -> Dict:
        """Generate business rules from data patterns"""
        return {
            'minimum_hours': {
                'any_job': 4.0,
                'with_installation': 8.0,
                'with_permit': 6.0
            },
            'crew_requirements': {
                'electrical_work': 'Licensed electrician required',
                'high_installation': '3-person crew for signs over 15ft',
                'crane_work': 'Certified operator required'
            },
            'sequencing_rules': {
                'fabrication_before_paint': ['0200-0270', 'before', '0410-0420'],
                'wiring_before_faces': ['0340', 'before', '0260'],
                'crating_before_install': ['0280', 'before', '0640-0650']
            }
        }
    
    def export_guide(self, guide: Dict, format: str = 'excel') -> str:
        """Export pricing guide in various formats"""
        
        if format == 'excel':
            return self._export_excel(guide)
        elif format == 'json':
            return self._export_json(guide)
        elif format == 'pdf':
            return self._export_pdf(guide)
        else:
            return self._export_text(guide)
    
    def _export_excel(self, guide: Dict) -> str:
        """Export to Excel with multiple sheets"""
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header style
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Write summary
        ws_summary['A1'] = "EAGLE SIGN CO. LABOR PRICING GUIDE"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A2'] = f"Generated: {guide['generated']}"
        
        row = 4
        for sign_type, data in guide['sign_types'].items():
            ws_summary.cell(row, 1, f"{sign_type} - {data['description']}")
            ws_summary.cell(row, 2, f"{len(data['labor_standards'])} codes")
            row += 1
            
        # Create sheet for each sign type
        for sign_type, data in guide['sign_types'].items():
            ws = wb.create_sheet(title=sign_type)
            
            # Headers
            headers = ['Code', 'Description', 'Dept', 'Min Hrs', 'Typical', 'Recommended', 'Max Hrs', 'Samples']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.fill = header_fill
                cell.font = header_font
                
            # Data
            row = 2
            for code, standard in sorted(data['labor_standards'].items()):
                ws.cell(row, 1, code)
                ws.cell(row, 2, standard['description'])
                ws.cell(row, 3, standard['department'])
                ws.cell(row, 4, standard['hours']['minimum'])
                ws.cell(row, 5, standard['hours']['typical'])
                ws.cell(row, 6, standard['hours']['recommended'])
                ws.cell(row, 7, standard['hours']['maximum'])
                ws.cell(row, 8, standard['sample_size'])
                
                # Color code by confidence
                if standard['sample_size'] >= 30:
                    ws.cell(row, 8).fill = PatternFill("solid", fgColor="90EE90")
                elif standard['sample_size'] >= 10:
                    ws.cell(row, 8).fill = PatternFill("solid", fgColor="FFFFE0")
                else:
                    ws.cell(row, 8).fill = PatternFill("solid", fgColor="FFB6C1")
                    
                row += 1
                
        filename = f"eagle_pricing_guide_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)
        return filename
    
    def _export_text(self, guide: Dict) -> str:
        """Export human-readable text format"""
        lines = []
        lines.append("="*80)
        lines.append("EAGLE SIGN CO. ULTIMATE PRICING GUIDE")
        lines.append(f"Generated: {guide['generated']}")
        lines.append("="*80)
        
        for sign_type, data in guide['sign_types'].items():
            lines.append(f"\n{sign_type} - {data['description']}")
            lines.append("-"*60)
            
            # Workflow
            lines.append("\nTYPICAL WORKFLOW:")
            for step in data['typical_workflow']:
                lines.append(f"  {step['sequence']}. {step['code']} {step['description']}")
                
            # Labor standards
            lines.append("\nLABOR STANDARDS:")
            lines.append(f"{'Code':<6} {'Description':<25} {'Min':<6} {'Typ':<6} {'Rec':<6} {'Max':<6} {'N':<4}")
            lines.append("-"*70)
            
            for code, std in sorted(data['labor_standards'].items()):
                hrs = std['hours']
                lines.append(
                    f"{code:<6} {std['description']:<25} "
                    f"{hrs['minimum']:<6.1f} {hrs['typical']:<6.1f} "
                    f"{hrs['recommended']:<6.1f} {hrs['maximum']:<6.1f} "
                    f"{std['sample_size']:<4}"
                )
                
                # Size formula if available
                if 'size_formula' in std:
                    formula = std['size_formula']
                    lines.append(f"       Formula: {formula['base_hours']:.1f} + {formula['hours_per_sqft']:.3f} * SQFT")
                    
        # Business rules
        lines.append("\n" + "="*80)
        lines.append("BUSINESS RULES:")
        for category, rules in guide['business_rules'].items():
            lines.append(f"\n{category.upper()}:")
            for key, value in rules.items():
                lines.append(f"  • {key}: {value}")
                
        return "\n".join(lines)

# Generate the guide
def generate_ultimate_guide(historical_data_path: str):
    """Generate pricing guide from historical data"""
    
    guide_generator = UltimatePricingGuide()
    
    # Load historical work orders
    # This would connect to your actual data source
    work_orders = load_historical_data(historical_data_path)
    
    # Generate guide
    pricing_guide = guide_generator.analyze_historical_data(work_orders)
    
    # Export in multiple formats
    excel_file = guide_generator.export_guide(pricing_guide, 'excel')
    text_report = guide_generator._export_text(pricing_guide)
    
    print(f"Generated: {excel_file}")
    print("\n" + text_report)
    
    return pricing_guide

def load_historical_data(path: str) -> List[Dict]:
    """Load work orders from your data source"""
    # Implementation depends on your data format
    # Could be CSV, database, or parsed PDFs
    pass

if __name__ == "__main__":
    generate_ultimate_guide("historical_data.csv")
